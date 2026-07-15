"""
Build data.json for the Oregon Elk Hunting Data Explorer static website.

Run this locally after you have new source files (or new years' worth of files):

    pip install pandas openpyxl pypdf
    python scripts/build_data.py --data-dir ./raw_data --out ./data/data.json

Then commit and push data/data.json (and the site files) to GitHub - GitHub Pages
serves it automatically, no server-side code needed. The webpage fetches
data/data.json at runtime and does all the plotting/report logic in the browser.

Filenames in --data-dir must contain the 4-digit hunting year, e.g.:
    2023_Elk_Preference_Point_Draw_Report.xlsx
    2023_Elk_Harvest_Summary.pdf
"""
import re
import glob
import os
import json
import argparse
from pathlib import Path
from collections import defaultdict

import numpy as np
import pandas as pd
import openpyxl
import pypdf

APP_YEARS = [2021, 2022, 2023, 2024, 2025, 2026]
HARVEST_YEARS = [2021, 2022, 2023, 2024, 2025]
DRAW_AVG_YEARS = [2024, 2025, 2026]

YEAR_RE = re.compile(r'(20\d{2})')


def _year_from_filename(path, valid_years):
    name = os.path.basename(path)
    for m in YEAR_RE.finditer(name):
        y = int(m.group(1))
        if y in valid_years:
            return y
    return None


def find_year_files(data_dir, pattern, valid_years):
    out = {}
    for f in glob.glob(str(Path(data_dir) / pattern)):
        y = _year_from_filename(f, valid_years)
        if y is not None:
            if y in out:
                print(f"WARNING: multiple files matched year {y}: {out[y]} and {f}. Keeping first.")
            else:
                out[y] = f
    return out


def parse_draw_report(path, year):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    header_idx = None
    for i, r in enumerate(rows):
        if r and r[0] == 'Hunt Number':
            header_idx = i
            break
    if header_idx is None:
        raise ValueError(f"Could not find 'Hunt Number' header row in {path}")

    records = []
    for r in rows[header_idx + 1:]:
        hunt_number = r[0]
        if not hunt_number or not isinstance(hunt_number, str):
            continue
        hunt_number = hunt_number.strip()
        if not hunt_number:
            continue
        try:
            records.append({
                'year': year,
                'hunt_number': hunt_number,
                'hunt_name': r[1],
                'tags_authorized': r[2],
                'resident_apps': r[3] or 0,
                'resident_drawn': r[4] or 0,
                'nonres_apps': r[5] or 0,
                'nonres_drawn': r[6] or 0,
                'total_apps': r[7] or 0,
                'total_drawn': r[8] or 0,
            })
        except IndexError:
            continue
    return pd.DataFrame.from_records(records)


def load_all_draw_data(data_dir, filename_pattern="*.xlsx"):
    files = find_year_files(data_dir, filename_pattern, APP_YEARS)
    missing = sorted(set(APP_YEARS) - set(files))
    if missing:
        print(f"NOTE: no draw-report file found for application year(s): {missing}")
    frames = [parse_draw_report(f, y) for y, f in sorted(files.items())]
    if not frames:
        raise ValueError("No draw report data loaded - check --data-dir and filenames.")
    return pd.concat(frames, ignore_index=True)


SUCCESS_RE = re.compile(r'(\d+(?:\.\d+)?)\s*%\s*$')
DIGIT_RE = re.compile(r'\d')
NUM_RE = re.compile(r'\b\d+\b')


def parse_harvest_pdf(path, year, known_hunt_numbers_sorted):
    """See notebook for full rationale. Summary: hunt number located by first digit in the
    line + longest-known-string match; the 11 numeric columns preceding 'NN%' are Total
    Hunters, Days Hunted, Antlerless, Total Bulls, Harvest Total, Spike, 2pt, 3pt, 4pt, 5pt,
    6+pt. When a hunt code appears on more than one row in the same report, hunters and
    harvest are summed and success % is *recomputed* from the summed totals (not averaged
    from the printed percentages, which can be wildly skewed by small-sample rows)."""
    reader = pypdf.PdfReader(path)
    text = "\n".join((p.extract_text() or "") for p in reader.pages)

    harvest_hits = defaultdict(list)
    hunters_hits = defaultdict(list)
    n_matches = defaultdict(int)

    for line in text.split("\n"):
        m = SUCCESS_RE.search(line)
        if not m:
            continue
        dm = DIGIT_RE.search(line)
        if not dm:
            continue
        idx = dm.start()
        hunt_number = None
        for cand in known_hunt_numbers_sorted:
            if line[idx:idx + len(cand)] == cand:
                hunt_number = cand
                break
        if not hunt_number:
            continue

        n_matches[hunt_number] += 1

        nums = NUM_RE.findall(line[:m.start()])
        if len(nums) >= 11:
            data_cols = [int(x) for x in nums[-11:]]
            hunters_hits[hunt_number].append(data_cols[0])
            harvest_hits[hunt_number].append(data_cols[4])

    records = []
    for hn, count in n_matches.items():
        h_vals = harvest_hits.get(hn, [])
        hunter_vals = hunters_hits.get(hn, [])
        hunters_sum = int(np.sum(hunter_vals)) if hunter_vals else None
        harvest_sum = int(np.sum(h_vals)) if h_vals else None
        success_pct = (harvest_sum / hunters_sum * 100) if hunters_sum else None
        records.append({
            'year': year,
            'hunt_number': hn,
            'success_pct': success_pct,
            'harvest_total': harvest_sum,
            'hunters': hunters_sum,
            'n_matches': count,
        })
    return pd.DataFrame.from_records(records)


def load_all_harvest_data(data_dir, known_hunt_numbers, filename_pattern="*.pdf"):
    known_sorted = sorted(set(known_hunt_numbers), key=len, reverse=True)
    files = find_year_files(data_dir, filename_pattern, HARVEST_YEARS)
    missing = sorted(set(HARVEST_YEARS) - set(files))
    if missing:
        print(f"NOTE: no harvest-summary file found for harvest year(s): {missing}")
    frames = [parse_harvest_pdf(f, y, known_sorted) for y, f in sorted(files.items())]
    if not frames:
        raise ValueError("No harvest data loaded - check --data-dir and filenames.")
    return pd.concat(frames, ignore_index=True)


def elk_200_series_hunts(draw_df):
    hunts = sorted(draw_df['hunt_number'].unique())
    return [h for h in hunts if re.match(r'^2\d{2}', h)]


def draw_pct_3yr_avg(draw_df, avg_years=DRAW_AVG_YEARS):
    sub = draw_df[draw_df.year.isin(avg_years)].copy()
    sub['draw_pct'] = np.where(sub.total_apps > 0, sub.total_drawn / sub.total_apps * 100, np.nan)
    return sub.groupby('hunt_number')['draw_pct'].mean().reset_index()


def success_5yr_stats(harvest_df):
    g = harvest_df.groupby('hunt_number')['success_pct'].agg(['mean', 'std', 'count']).reset_index()
    return g.rename(columns={'mean': 'success_mean', 'std': 'success_std', 'count': 'n_years'})


def scatter_data(draw_df, harvest_df, hunt_list):
    d = draw_pct_3yr_avg(draw_df)
    s = success_5yr_stats(harvest_df)
    merged = d.merge(s, on='hunt_number', how='inner')
    return merged[merged.hunt_number.isin(hunt_list)]


def _clean(v):
    """Convert numpy/pandas scalars to plain JSON-safe python values."""
    if v is None:
        return None
    if isinstance(v, (np.integer,)):
        return int(v)
    if isinstance(v, (np.floating,)):
        return None if np.isnan(v) else float(v)
    if isinstance(v, float) and np.isnan(v):
        return None
    return v


def build_json(data_dir):
    draw_df = load_all_draw_data(data_dir)
    known_hunt_numbers = draw_df['hunt_number'].unique().tolist()
    harvest_df = load_all_harvest_data(data_dir, known_hunt_numbers)
    elk_hunts = elk_200_series_hunts(draw_df)
    sd_all = scatter_data(draw_df, harvest_df, elk_hunts)

    dupe_flags = harvest_df[harvest_df.n_matches > 1]
    if len(dupe_flags):
        print(f"NOTE: {len(dupe_flags)} hunt/year combos had multiple matching PDF rows "
              f"(hunters/harvest summed, success % recomputed from the totals).")

    hunts = {}
    for hn in sorted(draw_df.hunt_number.unique()):
        draw_rows = draw_df[draw_df.hunt_number == hn].sort_values('year')
        harvest_rows = harvest_df[harvest_df.hunt_number == hn].sort_values('year')
        hunt_name = None
        if draw_rows.hunt_name.notna().any():
            hunt_name = draw_rows.hunt_name.dropna().iloc[-1]

        draw_list = []
        for _, r in draw_rows.iterrows():
            draw_list.append({
                'year': _clean(r.year),
                'total_apps': _clean(r.total_apps),
                'tags_authorized': _clean(r.tags_authorized),
                'resident_apps': _clean(r.resident_apps),
                'resident_drawn': _clean(r.resident_drawn),
                'nonres_apps': _clean(r.nonres_apps),
                'nonres_drawn': _clean(r.nonres_drawn),
                'total_drawn': _clean(r.total_drawn),
            })

        harvest_list = []
        for _, r in harvest_rows.iterrows():
            harvest_list.append({
                'year': _clean(r.year),
                'hunters': _clean(r.hunters),
                'harvest_total': _clean(r.harvest_total),
                'success_pct': _clean(r.success_pct),
            })

        hunts[hn] = {
            'name': hunt_name,
            'draw': draw_list,
            'harvest': harvest_list,
        }

    scatter = []
    for _, r in sd_all.iterrows():
        scatter.append({
            'hunt_number': r.hunt_number,
            'draw_pct': _clean(r.draw_pct),
            'success_mean': _clean(r.success_mean),
            'success_std': _clean(r.success_std),
            'n_years': _clean(r.n_years),
        })

    return {
        'meta': {
            'app_years': APP_YEARS,
            'harvest_years': HARVEST_YEARS,
            'draw_avg_years': DRAW_AVG_YEARS,
        },
        'elk_hunts': elk_hunts,
        'hunts': hunts,
        'scatter': scatter,
    }


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--data-dir", default="./raw_data", help="Folder with the source .xlsx/.pdf files")
    ap.add_argument("--out", default="./data/data.json", help="Output JSON path")
    args = ap.parse_args()

    payload = build_json(args.data_dir)
    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w") as f:
        json.dump(payload, f, separators=(",", ":"))
    print(f"Wrote {out_path} ({out_path.stat().st_size / 1024:.1f} KB), "
          f"{len(payload['hunts'])} hunts, {len(payload['elk_hunts'])} 200-series elk hunts.")
